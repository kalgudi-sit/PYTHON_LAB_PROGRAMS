#cell 1:
import numpy as np
import pandas as pd
import openpyxl as xl
from openpyxl import workbook

#cell 2:
df = pd.read_excel("E:\\NT\propertyData.xlsx")
x = df.values
x

#cell 3:
while 1:
    PID = int(input("\nEnter the PID : "))
    
    pos = -1

    for i in range(len(x)):
        if x[i][1] == PID:
            pos = i
            break
            
    if pos == -1:
        print("\nInvalid PID!!!! Please enter correct PID (301 - 310)")
        continue;
    else:
        print("\nPID",+PID,"exists at index : ", i+1)
                
        if x[i][4]<=0:
            print("\nNO TAX DUE")
        else:
            print("\nThe current dues are : ",x[i][4])
            TAX = int(input("\nEnte the tax amnt : "))
            print("\nAmount now paid is : ",TAX)
            balanceTax = x[i][4]-TAX
            if balanceTax>0:
                print("\nThe balance due is : ",balanceTax)
            elif balanceTax<0:
                print("\nThe Over paid balance is : ",balanceTax)
            else:
                print("\nNO TAX DUE")
            break;

wb = xl.load_workbook("E:\\NT\propertyData.xlsx")
sh = wb.active
sh.cell(row=i+2, column=6).value = TAX
sh.cell(row=i+2, column=5).value = balanceTax
wb.save("E:\\NT\propertyData.xlsx")  


